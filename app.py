from pathlib import Path

from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, flash
import os
import csv
from openpyxl import load_workbook
# import pandas as pd  # Временно отключено для тестирования
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, timedelta
import random
# from PIL import Image, ImageDraw  # Временно отключено для тестирования
from config import Config

# Initialize Flask application
app = Flask(__name__)
app.config.from_object(Config)

# Create data directory if it doesn't exist
os.makedirs(app.config['DATA_FILES']['crypto_images'], exist_ok=True)


# Context processor to inject current year into templates
@app.context_processor
def inject_now():
    return {
        'current_year': datetime.now().year,
        'now': datetime.now()
    }


def read_csv_file(file_path):
    """Читает CSV файл без pandas"""
    try:
        data = []
        encodings = ['utf-8', 'cp1251', 'latin-1', 'iso-8859-1', 'windows-1251']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as file:
                    reader = csv.reader(file, delimiter=';')
                    headers = next(reader)  # Первая строка - заголовки
                    for row in reader:
                        if len(row) == len(headers):  # Проверяем, что количество колонок совпадает
                            data.append(row)
                return headers, data
            except UnicodeDecodeError:
                continue
        
        raise Exception("Не удалось прочитать файл с любой из попробованных кодировок")
    except Exception as e:
        raise Exception(f"Ошибка чтения CSV файла: {str(e)}")


def read_excel_file(file_path):
    """Читает Excel файл без pandas"""
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        
        data = []
        headers = []
        
        for row in sheet.iter_rows(values_only=True):
            if not headers:
                headers = [str(cell) if cell is not None else '' for cell in row]
            else:
                data.append([str(cell) if cell is not None else '' for cell in row])
        
        workbook.close()
        return headers, data
    except Exception as e:
        raise Exception(f"Ошибка чтения Excel файла: {str(e)}")


def generate_html_table(headers, data):
    """Генерирует HTML таблицу из данных"""
    if not headers or not data:
        return "<p>Нет данных для отображения</p>"
    
    html = '<table class="table table-striped table-hover">'
    
    # Заголовки
    html += '<thead><tr>'
    for header in headers:
        html += f'<th>{header}</th>'
    html += '</tr></thead>'
    
    # Данные
    html += '<tbody>'
    for row in data:
        html += '<tr>'
        for cell in row:
            html += f'<td>{cell}</td>'
        html += '</tr>'
    html += '</tbody>'
    
    html += '</table>'
    return html


# Generate mock data if files don't exist
if app.config['MOCK_DATA']:
    for file_type, file_path in app.config['DATA_FILES'].items():
        if file_type != 'crypto_images' and not os.path.exists(file_path):
            if file_type == 'clients':
                # Generate client data - используем CSV модуль
                try:
                    client_ids = list(range(127051970, 127051970 + 100))
                    usernames = [f'user_{num}' for num in range(1, 101)]
                    ips = [f'192.168.{random.randint(0, 255)}.{random.randint(0, 255)}' for _ in range(100)]
                    start_dates = [(datetime.now() - timedelta(days=random.randint(0, 30))).strftime('%d.%m.%Y %H:%M') for _ in range(100)]
                    end_dates = [(datetime.now() - timedelta(days=random.randint(0, 30))).strftime('%d.%m.%Y %H:%M') for _ in range(100)]

                    os.makedirs(os.path.dirname(file_path), exist_ok=True)
                    with open(file_path, 'w', newline='', encoding='utf-8') as file:
                        writer = csv.writer(file, delimiter=';')
                        writer.writerow(['ID', 'Username', 'IP', 'Start date', 'End date'])
                        for i in range(100):
                            writer.writerow([client_ids[i], usernames[i], ips[i], start_dates[i], end_dates[i]])
                    print(f"Generated mock CSV file: {file_path}")
                except Exception as e:
                    print(f"Error generating mock CSV: {e}")

            elif file_type == 'hot_leads':
                # Generate hot leads data - используем openpyxl
                try:
                    from openpyxl import Workbook
                    workbook = Workbook()
                    sheet = workbook.active
                    
                    # Заголовки
                    headers = ['№', 'User', 'IP', 'Report', 'Start date', 'End date', 'Картинка']
                    for col, header in enumerate(headers, 1):
                        sheet.cell(row=1, column=col, value=header)
                    
                    # Данные
                    for row in range(2, 22):  # 20 строк данных
                        sheet.cell(row=row, column=1, value=126922958 + row - 2)
                        sheet.cell(row=row, column=2, value=f'user_{row-1}')
                        sheet.cell(row=row, column=3, value=f'10.0.{random.randint(0, 255)}.{random.randint(0, 255)}')
                        sheet.cell(row=row, column=4, value=f'Sample report {row-1}')
                        sheet.cell(row=row, column=5, value=(datetime.now() - timedelta(days=random.randint(0, 30))).strftime('%d.%m.%Y %H:%M'))
                        sheet.cell(row=row, column=6, value=(datetime.now() - timedelta(days=random.randint(0, 30))).strftime('%d.%m.%Y %H:%M'))
                        sheet.cell(row=row, column=7, value=f'/uploads/files/20250606/{random.randint(1000000000, 9999999999)}.png')
                    
                    os.makedirs(os.path.dirname(file_path), exist_ok=True)
                    workbook.save(file_path)
                    workbook.close()
                    print(f"Generated mock Excel file: {file_path}")
                except Exception as e:
                    print(f"Error generating mock Excel: {e}")

    # Create mock crypto images - временно отключено Pillow
    crypto_dir = app.config['DATA_FILES']['crypto_images']
    if not os.listdir(crypto_dir):
        print("Mock image generation temporarily disabled")
        pass

# User credentials for authentication
users = {
    'admin': generate_password_hash('admin123'),
    'manager': generate_password_hash('manager123')
}


# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Please log in to access this page', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)

    return decorated_function


# Application routes
@app.route('/')
def home():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username in users and check_password_hash(users[username], password):
            session['username'] = username
            next_page = request.args.get('next')
            flash('Login successful!', 'success')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        flash('Invalid username or password', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('username', None)
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')


@app.route('/clients')
@login_required
def show_clients():
    csv_path = app.config['DATA_FILES'].get('clients', 'data/finish ltt.csv')
    try:
        headers, data = read_csv_file(csv_path)
        table_html = generate_html_table(headers, data)
        return render_template('clients.html', table=table_html, error=None)
    except Exception as e:
        flash(f'Ошибка при загрузке данных клиентов: {str(e)}', 'danger')
        return render_template('clients.html', table=None, error=str(e))


@app.route('/hot-leads')
@login_required
def show_hot_leads():
    excel_path = app.config['DATA_FILES'].get('hot_leads', 'data/filtered_69_images_HQ_fast.xlsx')
    try:
        headers, data = read_excel_file(excel_path)
        table_html = generate_html_table(headers, data)
        return render_template('hot_lead.html', table=table_html, error=None)
    except Exception as e:
        flash(f'Ошибка при загрузке Hot Lead данных: {str(e)}', 'danger')
        return render_template('hot_lead.html', table=None, error=str(e))


@app.route('/crypto-report')
def show_crypto_report():
    try:
        image_dir = app.config['DATA_FILES']['crypto_images']
        images = [f for f in os.listdir(image_dir)
                  if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif'))]

        images.sort()
        return render_template('crypto_report.html', images=images)
    except Exception as e:
        return f"Помилка завантаження зображень: {str(e)}", 500


@app.route('/crypto-images/<filename>')
def get_crypto_image(filename):
    return send_from_directory(
        app.config['DATA_FILES']['crypto_images'],
        filename
    )


@app.route('/crypto-images/<filename>')
@login_required
def serve_crypto_image(filename):
    return send_from_directory(app.config['DATA_FILES']['crypto_images'], filename)


def analyze_images(image_folder='static/crypto-images'):
    images_data = []
    formats = set()

    for filename in os.listdir(image_folder):
        filepath = os.path.join(image_folder, filename)

        if not os.path.isfile(filepath):
            continue

        try:
            with Image.open(filepath) as img:
                width, height = img.size
                fmt = img.format or 'Невідомий формат'
                formats.add(fmt)

            size_kb = round(os.path.getsize(filepath) / 1024, 2)
            modified_time = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%d.%m.%Y %H:%M')

            img_data = {
                'filename': filename,
                'width': width,
                'height': height,
                'size_kb': size_kb,
                'format': fmt,  # виправлено: тут має бути результат image.format, а не метод str.format
                'modified': modified_time
            }

            images_data.append(img_data)
        except Exception as e:
            print(f"Помилка при обробці зображення '{filename}': {e}")
            continue

    stats = {
        'total': len(images_data),
        'formats': sorted(formats),
        'last_update': datetime.now().strftime('%d.%m.%Y %H:%M')
    }

    return images_data, stats

@app.route('/crypto-report')
def crypto_report():
    image_dir = 'data/crypto-images'
    images = []
    for filename in os.listdir(image_dir):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
            path = os.path.join(image_dir, filename)
            with Image.open(path) as img:
                width, height = img.size
            images.append({
                'name': filename,
                'size': f"{round(os.path.getsize(path)/1024, 1)} KB",
                'dimensions': f"{width}×{height}",
                'format': filename.split('.')[-1].upper(),
                'modified': datetime.fromtimestamp(os.path.getmtime(path)).strftime('%d.%m.%Y %H:%M')
            })

    stats = {
        'total': len(images),
        'formats': sorted(set(img['format'] for img in images)),
        'last_update': datetime.now().strftime('%d.%m.%Y %H:%M')
    }

    return render_template('crypto_report.html', images=images, stats=stats)


@login_required  # або без, якщо не потрібна авторизація

@app.route('/crypto-images/<path:filename>')
def get_image(filename):
    return send_from_directory('static/crypto-images', filename)


@app.route('/images/<filename>')
def serve_image(filename):
    return send_from_directory(IMAGE_DIR, filename)

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(STATIC_DIR, filename)

# Налаштування шляхів
BASE_DIR = Path(__file__).parent
IMAGE_DIR = BASE_DIR / 'data' / 'crypto-images'
IMAGE_DIR.mkdir(parents=True, exist_ok=True)
STATIC_DIR = BASE_DIR / 'static'
STATIC_DIR.mkdir(exist_ok=True)

def get_image_data():
    """Отримує дані про всі зображення у папці"""
    images = []
    for img_file in IMAGE_DIR.glob('*'):
        if img_file.suffix.lower() in ('.png', '.jpg', '.jpeg', '.gif', '.webp'):
            try:
                with Image.open(img_file) as img:
                    images.append({
                        'name': img_file.name,
                        'size': f"{os.path.getsize(img_file)/1024:.1f} KB",
                        'width': img.width,
                        'height': img.height,
                        'format': img.format,
                        'modified': datetime.fromtimestamp(
                            os.path.getmtime(img_file)).strftime('%d.%m.%Y %H:%M')
                    })
            except Exception as e:
                print(f"Помилка обробки {img_file.name}: {e}")
    return sorted(images, key=lambda x: x['name'])

@app.route('/')
def index():
    images, stats = analyze_images()
    now = datetime.now()
    return render_template('index.html', images=images, stats=stats, now=now)

IMAGE_FOLDER = 'static/images'

def get_images():
    images = []
    for filename in os.listdir(IMAGE_FOLDER):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')):
            path = os.path.join(IMAGE_FOLDER, filename)
            stat = os.stat(path)
            size_kb = round(stat.st_size / 1024, 2)
            modified = datetime.fromtimestamp(stat.st_mtime).strftime('%d.%m.%Y %H:%M')
            # Для розмірів картинки можна використовувати PIL (Pillow)
            try:
                from PIL import Image
                with Image.open(path) as img:
                    width, height = img.size
                    format = img.format
            except ImportError:
                width = height = format = 'N/A'
            images.append({
                'filename': filename,
                'size_kb': size_kb,
                'modified': modified,
                'width': width,
                'height': height,
                'format': format
            })
    return images

@app.route('/gallery')
def gallery():
    images = get_images()
    stats = {
        'total': len(images),
        'formats': list(set(img['format'] for img in images if img['format'] != 'N/A'))
    }
    now = datetime.now()
    return render_template('gallery.html', images=images, stats=stats, now=now)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)